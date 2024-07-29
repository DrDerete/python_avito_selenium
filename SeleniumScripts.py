from selenium import webdriver
from selenium.common import exceptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook, load_workbook
from os.path import exists
from itertools import islice
import urllib.request as req
import WebAdress as Web
import numpy as np
import time
import cv2


class Driver(webdriver.Chrome):

    def __init__(self, service=webdriver.ChromeService(), options=webdriver.ChromeOptions()):  # настройки для selenium.webdriver
        self.__last = False
        self.__next_page = None
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        super().__init__(options, service)
        self.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    def __solve_captcha(self):  # тестил прохождение капчи
        while True:
            WebDriverWait(self, timeout=5).until(lambda x: x.find_element("xpath", "//button")).click()
            distance = self.__parse_image()
            slider = self.find_element("xpath", "//div[contains(@class, \"geetest_btn\")]")
            actions = webdriver.ActionChains(driver=self)
            actions.pause(3).perform()
            actions.click_and_hold(slider).pause(1.732).perform()
            actions.move_by_offset(distance, 0).pause(1.321).perform()
            actions.release().pause(3).perform()
            time.sleep(12)
            if "IP" not in self.title:
                return

    def __parse_image(self):  # скачивание изображения (из кеша драйвера не даёт)
        web_elements = WebDriverWait(self, timeout=5).until(lambda x: x.find_elements("xpath", "//div[contains(@class, \"window\")]/div"))
        back = web_elements[1].get_attribute("style").split("\"")[1]
        req.urlretrieve(back, "Tota/back.png")
        slide = web_elements[0].find_element("xpath", "//div[contains(@class, \"slice_bg\")]").get_attribute("style").split("\"")[1]
        req.urlretrieve(slide, "Tota/slide.png")
        return self.__find_distance()

    def __find_distance(self):  # найти дистанцию сдвига слайдера на капче
        slide_info = self.__what_about()
        middle_slide = slide_info["top"] + int(slide_info["width"] / 2)
        origin = self.__count_slide(int(slide_info["width"] / 2))
        image_back = cv2.imread("Tota/back.png")
        gray_image = cv2.cvtColor(image_back, cv2.COLOR_BGR2GRAY)
        pixels = []
        loop = iter(range(slide_info["width"], len(gray_image[middle_slide])-int(slide_info["width"]/2)))
        for i in loop:
            if gray_image[middle_slide][i] / gray_image[middle_slide][i + 1] > 1.35:  # окрашимаем изображение в 2-color для упращения и ищем паттерн затемения, добовляем все варианты
                pixels.append(i)
                next(islice(loop, 9, 10), None)
        match len(pixels):  # смотрим сколько, резко затемненных пикселей было найдено
            case 0:
                print("Багуля")
                return 0
            case 1:
                return pixels[0] - int(origin["zeros"]/2)  # если один, то сразу наш клиент
            case _: # в противном случае смотрим сумму последующих пикселей, как у слайдера
                pix_area = []
                for i in pixels:
                    pix_area.append(gray_image[middle_slide][i:i + origin["colors"]])
                indicator = 0
                index = 0
                for i in range(len(pix_area)): 
                    factor = sum(pix_area[i]) / sum(origin["pattern"])
                    if abs(factor - 0.5) < abs(indicator - 0.5): # вот такое получиось, хотя думаю можно поэлегантней придумать условие
                        indicator = factor
                        index = i
                return pixels[index] - int(origin["zeros"] / 2)

    def __what_about(self):
        slide = {}
        info = ":".join(
            self.find_element("xpath", "(//div[contains(@class, \"slice\")])[1]").get_attribute("style").split(";")
        ).split(":")
        for i in range(0, 5, 2):
            slide[info[i].strip()] = info[i + 1]
        for key in slide:
            slide[key] = round(float(slide[key].split("p")[0]))
        return slide

    @staticmethod
    def __count_slide(middle):
        count_info = {"colors": 0, "zeros": 0, "pattern": []}
        image_slide = cv2.imread("Tota/slide.png")
        gray_image = cv2.cvtColor(image_slide, cv2.COLOR_BGR2GRAY)
        for pixel_index in range(len(gray_image[middle])):
            if gray_image[middle][pixel_index] != 0:
                count_info["colors"] += 1
                count_info["pattern"].append(gray_image[middle][pixel_index])
            else:
                count_info["zeros"] += 1
        return count_info

    @staticmethod
    def __hand_info(string):
        info = []
        mas = string.split("\n")
        match (len(mas)):
            case 2:
                info.append(f"Organization:Unknown")
            case 3:
                info.append(f"Organization:{mas[0]}")
                info.append(f"Based:{mas[2]}")
            case 4:
                if mas[1] == "Работодатель":
                    info.append(f"Organization:{mas[0]}")
                    info.append(f"Based:{mas[2]}")
                else:
                    info.append(f"Organization:{mas[0]}")
                    info.append(f"Score:{mas[1]}")
                    info.append(f"Comment:{mas[2]}")
            case 5:
                info.append(f"Organization:{mas[0]}")
                info.append(f"Score:{mas[1]}")
                info.append(f"Comment:{mas[2]}")
                info.append(f"Based:{mas[4]}")
        return "?".join(info)

    def __get(self, uri):
        self.get(uri)
        if "IP" in self.title:
            self.__solve_captcha()

    def __new_page(self, elem):
        elem.click()
        WebDriverWait(self, timeout=5).until(EC.number_of_windows_to_be(2))
        self.switch_to.window(self.window_handles[1])
        if "IP" in self.title:
            self.__solve_captcha()

    def __exist_next_page(self):
        try:
            return self.find_element(*Web.NEXT_PAGE).get_attribute("href")
        except exceptions.NoSuchElementException:
            return "END"

    def parsing(self, uri, expended, vacancy): 
        if expended:
            pass
            # тут составить url из ввода
        self.__get(uri)
        data = {}
        original_window = self.current_window_handle
        works = self.find_elements(*Web.WORKS)
        for elem in works:
            self.__new_page(elem)
            work = self.find_element(*Web.WORK_TITLE).text
            price = self.find_element(*Web.PRICE).text
            conditions = self.find_element(*Web.CONDITIONS).text
            location = self.find_element(*Web.LOCATION).text
            description = self.find_element(*Web.DESCRIPTION).text
            info = self.find_element(*Web.INFO).text
            data[elem] = [work, price, conditions, location, description, self.__hand_info(info)]
            self.close()
            self.switch_to.window(original_window)
        Excell().save(data)  # в зависимости от запроса сделать ограничение по количеству
        return self.__exist_next_page()


class Excell(object):  # сохранение в ex

    path = "Tota/Avito.xlsx"

    def save(self, data):
        if not exists(self.path):
            self.create_table()
        wb = load_workbook(self.path)
        sheet = wb.active
        for key in data:
            row = str(sheet.max_row + 1)
            for ch in "ABCDEF":
                sheet[ch + row] = data[key]["ABCDEF".index(ch)]
        wb.save(self.path)

    def create_table(self):
        wb = Workbook()
        wb.active.title = "Avito"
        wb.active["A1"] = "Work"
        wb.active["B1"] = "Price"
        wb.active["C1"] = "Conditions"
        wb.active["D1"] = "Locations"
        wb.active["E1"] = "Description"
        wb.active["F1"] = "Info"
        wb.save(self.path)
